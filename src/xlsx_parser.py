#!/usr/bin/env python3
"""
計画.xlsx パーサー
縦断・横断データをJSON形式で出力する

Geminiを使う場合は環境変数 GEMINI_API_KEY を設定する。
--local-only でローカル解析に固定できる。
"""

import json
import math
import os
import re
import sys
import urllib.error
import urllib.request
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Optional

import openpyxl


@dataclass
class SurveyRow:
    """横断の1点データ"""
    unit_distance: float       # 単延長（間隔）
    elevation: float           # 現地盤高
    planned_height: float      # 計画高
    cumulative_distance: float # 累計距離（CLからの距離、左がマイナス）
    cutting_bottom: float      # 切削底面高


@dataclass
class CrossSectionData:
    """横断データ"""
    survey_point_name: str        # 測点名
    dl: float                     # DL（基準線高さ）
    cl_index: int                 # センターラインのインデックス
    l_to_cl_distance: float       # 左端からCLまでの距離
    survey_data: list[SurveyRow]  # 測量データ
    route_distance: Optional[float] = None  # 路線距離（m）
    route_id: str = "route_1"     # ルートID


@dataclass
class LongitudinalPoint:
    """縦断の1点データ"""
    station_name: str      # 測点名
    ground_height: float   # 地盤高（現況）
    planned_height: float  # 計画高
    distance: float        # 追加距離
    cumulative: float      # 累積距離
    gradient: float        # 勾配


@dataclass
class LongitudinalSection:
    """縦断データ"""
    name: str                      # 路線名
    points: list[LongitudinalPoint]


def parse_station_distance(name: str) -> float:
    """測点名から路線距離を計算する
    例: No.0 -> 0, No.1 -> 20, No.1+10 -> 30, No.2+5.5 -> 45.5
    """
    # "No.X" or "No.X+Y" or "No.X+Yマンホール" のパターン
    match = re.match(r'No\.(\d+)(?:\+(\d+(?:\.\d+)?))?\s*', name)
    if not match:
        return 0.0

    no = int(match.group(1))
    plus = float(match.group(2)) if match.group(2) else 0.0
    # No.1 = 20m間隔と仮定
    return no * 20.0 + plus


def sheet_to_table(ws) -> list[dict]:
    """シートを行配列に変換（空行は除外）"""
    max_row = 0
    max_col = 0
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v is not None and str(v).strip() != "" for v in row):
            max_row = r_idx
            for c_idx, val in enumerate(row, start=1):
                if val is not None and str(val).strip() != "":
                    max_col = max(max_col, c_idx)

    if max_row == 0 or max_col == 0:
        return []

    rows = []
    for r_idx in range(1, max_row + 1):
        values = []
        for c_idx in range(1, max_col + 1):
            val = ws.cell(row=r_idx, column=c_idx).value
            values.append("" if val is None else str(val))
        if any(v != "" for v in values):
            rows.append({"row": r_idx, "values": values})
    return rows


def workbook_to_prompt_data(workbook) -> dict:
    """Gemini用の入力データを作成"""
    sheets = []
    for ws in workbook.worksheets:
        sheets.append({
            "name": ws.title,
            "rows": sheet_to_table(ws),
        })
    return {"sheets": sheets}


def extract_json_from_text(text: str) -> str:
    """コードフェンスを除去してJSON文字列を抽出"""
    if "```" not in text:
        return text.strip()
    parts = text.split("```")
    if len(parts) >= 2:
        candidate = parts[1]
        # 先頭の言語指定を削除
        if candidate.startswith("json"):
            candidate = candidate[4:]
        return candidate.strip()
    return text.strip()


def call_gemini(prompt: str, model: str) -> str:
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("GEMINI_API_KEY is not set")

    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.2,
            "topP": 0.9,
            "maxOutputTokens": 8192,
        },
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"Gemini API error: {e.code} {e.reason}") from e

    result = json.loads(body)
    candidates = result.get("candidates", [])
    if not candidates:
        raise RuntimeError("Gemini API returned no candidates")
    parts = candidates[0].get("content", {}).get("parts", [])
    if not parts:
        raise RuntimeError("Gemini API returned no content parts")
    return parts[0].get("text", "")


def parse_xlsx_with_gemini(xlsx_path: Path, model: str) -> list[dict]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    prompt_data = workbook_to_prompt_data(workbook)
    prompt = (
        "You are given an Excel workbook as JSON (sheets with rows and values).\n"
        "Extract road cross-section data and output ONLY a JSON array matching this schema:\n"
        "[{\n"
        "  \"survey_point_name\": string,\n"
        "  \"dl\": number,\n"
        "  \"cl_index\": number,\n"
        "  \"l_to_cl_distance\": number,\n"
        "  \"survey_data\": [{\n"
        "    \"unit_distance\": number,\n"
        "    \"elevation\": number,\n"
        "    \"planned_height\": number,\n"
        "    \"cumulative_distance\": number,\n"
        "    \"cutting_bottom\": number\n"
        "  }],\n"
        "  \"route_distance\": number | null,\n"
        "  \"route_id\": string\n"
        "}]\n"
        "Rules:\n"
        "- Output JSON only (no markdown).\n"
        "- If multiple routes exist, set route_id per route; otherwise use \"route_1\".\n"
        "- cl_index is the index of the center line point within survey_data.\n"
        "- l_to_cl_distance is the distance from left edge to CL (meters).\n"
        "- cumulative_distance is distance from CL (left is negative, right positive).\n"
        "- cutting_bottom is ground elevation minus cutting depth if explicit data is not present.\n"
        "Input workbook:\n"
    )
    prompt += json.dumps(prompt_data, ensure_ascii=False)

    text = call_gemini(prompt, model)
    json_text = extract_json_from_text(text)
    sections = json.loads(json_text)
    if not isinstance(sections, list):
        raise ValueError("Gemini output is not a JSON array")
    return sections


def parse_longitudinal_sheet(ws, sheet_name: str) -> LongitudinalSection:
    """縦断シートをパースする"""
    points = []
    cumulative = 0.0

    # 行3から開始（行2がヘッダー）
    for row_idx in range(3, ws.max_row + 1):
        station_name = ws.cell(row=row_idx, column=2).value
        if not station_name:
            continue

        ground_height = ws.cell(row=row_idx, column=3).value
        if ground_height is None:
            continue

        distance = ws.cell(row=row_idx, column=4).value or 0.0
        gradient = ws.cell(row=row_idx, column=5).value or 0.0

        cumulative += distance

        # 計画高は縦断(計画)シートから取得するか、現況と同じにする
        planned_height = ground_height  # 後で上書き

        points.append(LongitudinalPoint(
            station_name=str(station_name),
            ground_height=float(ground_height),
            planned_height=float(planned_height),
            distance=float(distance),
            cumulative=cumulative,
            gradient=float(gradient) if gradient else 0.0,
        ))

    return LongitudinalSection(name=sheet_name, points=points)


def parse_cross_section_sheet(ws, sheet_name: str, cutting_depth: float = 0.05) -> list[CrossSectionData]:
    """横断シートをパースする

    Excel構造:
    - col2 = 測点名
    - col3 = L標高（左端）
    - col4 = C標高（センター）
    - col5 = マンホール標高（オプション）
    - col6 = R標高（右端）
    - col8 = L距離（CLからの左幅）
    - col9 = R距離（CLからの右幅）
    - col10 = マンホール距離（オプション）
    """
    sections = []

    # 前回の距離を保持（欠損時のフォールバック用）
    prev_l_dist = 3.0
    prev_r_dist = 3.0

    # 行5から開始、2行ごとに1測点
    row_idx = 5
    while row_idx <= ws.max_row:
        station_name = ws.cell(row=row_idx, column=2).value
        if not station_name:
            row_idx += 2
            continue

        # 標高データを取得
        l_elev = ws.cell(row=row_idx, column=3).value   # L（左端）
        c_elev = ws.cell(row=row_idx, column=4).value   # C（センター）
        mh_elev = ws.cell(row=row_idx, column=5).value  # マンホール（オプション）
        r_elev = ws.cell(row=row_idx, column=6).value   # R（右端）

        # 距離データを取得
        l_dist_val = ws.cell(row=row_idx, column=8).value   # L距離
        r_dist_val = ws.cell(row=row_idx, column=9).value   # R距離
        mh_dist_val = ws.cell(row=row_idx, column=10).value # マンホール距離

        # 距離が欠損の場合は前回値を使用
        if l_dist_val is not None:
            l_dist = float(l_dist_val)
            prev_l_dist = l_dist
        else:
            l_dist = prev_l_dist

        if r_dist_val is not None:
            r_dist = float(r_dist_val)
            prev_r_dist = r_dist
        else:
            r_dist = prev_r_dist

        # データポイントを構築（左から右の順）
        points = []

        # L（左端）: 距離は負値
        if l_elev is not None:
            points.append((-l_dist, float(l_elev)))

        # C（センター）: 距離は0
        if c_elev is not None:
            points.append((0.0, float(c_elev)))

        # マンホール: 距離は正値
        if mh_elev is not None and mh_dist_val is not None:
            points.append((float(mh_dist_val), float(mh_elev)))

        # R（右端）: 距離は正値
        if r_elev is not None:
            points.append((r_dist, float(r_elev)))

        # 距離でソート
        points.sort(key=lambda p: p[0])

        if len(points) < 2:
            row_idx += 2
            continue

        # 距離と標高を分離
        distances = [p[0] for p in points]
        elevations = [p[1] for p in points]

        # CLのインデックスを見つける
        cl_index = 0
        for i, d in enumerate(distances):
            if abs(d) < 0.001:
                cl_index = i
                break

        # 単延長を計算
        unit_distances = [0.0]
        for i in range(1, len(distances)):
            unit_distances.append(distances[i] - distances[i-1])

        # 左端からCLまでの距離
        l_to_cl = abs(distances[0]) if distances else 0.0

        # DL（最小標高の小数点以下切り捨て）
        dl = math.floor(min(elevations))

        # SurveyRowを作成
        survey_data = []
        for i in range(len(elevations)):
            planned = elevations[i]  # 現況シートの場合は同じ値
            survey_data.append(SurveyRow(
                unit_distance=unit_distances[i],
                elevation=elevations[i],
                planned_height=planned,
                cumulative_distance=distances[i],
                cutting_bottom=planned - cutting_depth,
            ))

        route_distance = parse_station_distance(str(station_name))

        sections.append(CrossSectionData(
            survey_point_name=str(station_name),
            dl=dl,
            cl_index=cl_index,
            l_to_cl_distance=l_to_cl,
            survey_data=survey_data,
            route_distance=route_distance,
        ))

        row_idx += 2

    return sections


def parse_keikaku_matome_sheet(ws, cutting_depth: float = 0.05) -> list[CrossSectionData]:
    """計画まとめシートをパースする

    Excel構造（行1がヘッダー）:
    - col2 = 測点名（現地盤高の行のみ）
    - col3 = 行種別（現地盤高/計画高/切削高/切削厚）
    - col4 = L標高, col5 = -4m, col6 = -2m, col7 = C, col8 = +2m, col9 = +4m, col10 = +6m, col11 = +8m, col12 = R
    - col14 = 左幅員, col15 = 右幅員
    - col17 = 左勾配, col18 = 右勾配

    5行で1測点（現地盤高, 計画高, 切削高, 切削厚, 空行）
    空白行2つで次のルートに切り替わる
    """
    sections = []
    seen_names = set()  # 重複検出用（ルートごとにリセット）
    route_num = 1
    blank_count = 0  # 連続空白行カウント

    # 距離マッピング（col4-12の相対位置、後で実際の幅員で調整）
    # L=-幅員, -4, -2, C=0, +2, +4, +6, +8, R=+幅員
    dist_offsets = {
        4: None,   # L（左幅員で決まる）
        5: -4.0,
        6: -2.0,
        7: 0.0,    # C
        8: 2.0,
        9: 4.0,
        10: 6.0,
        11: 8.0,
        12: None,  # R（右幅員で決まる）
    }

    row_idx = 2  # 行2から開始（行1はヘッダー）
    while row_idx <= ws.max_row:
        station_name = ws.cell(row=row_idx, column=2).value
        row_type = ws.cell(row=row_idx, column=3).value

        # 空白行の検出
        if not station_name and not row_type:
            blank_count += 1
            row_idx += 1
            continue

        # 空白行2つ以上 + No.0で新ルート開始
        if blank_count >= 2 and station_name and "No.0" in str(station_name):
            route_num += 1
            seen_names = set()  # リセット

        blank_count = 0  # 空白行カウントリセット

        if not station_name or row_type != "現地盤高":
            row_idx += 1
            continue

        name_str = str(station_name)
        seen_names.add(name_str)

        # 幅員を取得
        l_width = ws.cell(row=row_idx, column=14).value
        r_width = ws.cell(row=row_idx, column=15).value

        if l_width is None or r_width is None:
            row_idx += 1
            continue

        l_width = float(l_width)
        r_width = float(r_width)

        # 現地盤高の行から標高を取得
        ground_elevations = {}
        for col in range(4, 13):
            val = ws.cell(row=row_idx, column=col).value
            if val is not None:
                ground_elevations[col] = float(val)

        # 計画高の行（次の行）から標高を取得
        planned_elevations = {}
        plan_row = row_idx + 1
        if ws.cell(row=plan_row, column=3).value == "計画高":
            for col in range(4, 13):
                val = ws.cell(row=plan_row, column=col).value
                if val is not None:
                    planned_elevations[col] = float(val)

        # データポイントを構築
        points = []
        for col, offset in dist_offsets.items():
            if col not in ground_elevations:
                continue

            # 距離を決定
            if col == 4:  # L
                dist = -l_width
            elif col == 12:  # R
                dist = r_width
            else:
                dist = offset
                # 幅員より外の点はスキップ
                if dist < -l_width or dist > r_width:
                    continue

            ground = ground_elevations[col]
            planned = planned_elevations.get(col, ground)

            points.append({
                'dist': dist,
                'ground': ground,
                'planned': planned,
            })

        # 距離でソート
        points.sort(key=lambda p: p['dist'])

        if len(points) < 2:
            row_idx += 1
            continue

        # CLのインデックスを見つける
        cl_index = 0
        for i, p in enumerate(points):
            if abs(p['dist']) < 0.001:
                cl_index = i
                break

        # 単延長を計算
        distances = [p['dist'] for p in points]
        unit_distances = [0.0]
        for i in range(1, len(distances)):
            unit_distances.append(distances[i] - distances[i-1])

        # DL（最小標高の小数点以下切り捨て）
        dl = math.floor(min(p['ground'] for p in points))

        # SurveyRowを作成
        survey_data = []
        for i, p in enumerate(points):
            survey_data.append(SurveyRow(
                unit_distance=unit_distances[i],
                elevation=p['ground'],
                planned_height=p['planned'],
                cumulative_distance=p['dist'],
                cutting_bottom=p['planned'] - cutting_depth,
            ))

        route_distance = parse_station_distance(str(station_name))

        sections.append(CrossSectionData(
            survey_point_name=str(station_name),
            dl=dl,
            cl_index=cl_index,
            l_to_cl_distance=l_width,
            survey_data=survey_data,
            route_distance=route_distance,
            route_id=f"route_{route_num}",
        ))

        row_idx += 1  # 次の行へ（5行ジャンプするとルート境界の空白行を見逃す）

    return sections


def parse_xlsx(xlsx_path: str) -> dict:
    """Excelファイルをパースして辞書形式で返す"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    result = {
        "longitudinal_current": None,
        "longitudinal_planned": None,
        "cross_sections_current": [],
        "cross_sections_planned": [],
    }

    # 縦断データ
    if "縦断 (現況)" in wb.sheetnames:
        result["longitudinal_current"] = asdict(
            parse_longitudinal_sheet(wb["縦断 (現況)"], "縦断(現況)")
        )

    if "縦断 (計画)" in wb.sheetnames:
        result["longitudinal_planned"] = asdict(
            parse_longitudinal_sheet(wb["縦断 (計画)"], "縦断(計画)")
        )

    # 横断データ - 計画まとめシートを優先使用
    if "計画まとめ" in wb.sheetnames:
        sections = parse_keikaku_matome_sheet(wb["計画まとめ"])
        result["cross_sections_merged"] = [asdict(s) for s in sections]
    else:
        # フォールバック: 横断(現況)と横断(計画)を使用
        if "横断(現況)" in wb.sheetnames:
            sections = parse_cross_section_sheet(wb["横断(現況)"], "横断(現況)")
            result["cross_sections_current"] = [asdict(s) for s in sections]

        if "横断(計画)" in wb.sheetnames:
            sections = parse_cross_section_sheet(wb["横断(計画)"], "横断(計画)")
            result["cross_sections_planned"] = [asdict(s) for s in sections]

        # 現況と計画をマージ（計画高を上書き）
        if result["cross_sections_current"] and result["cross_sections_planned"]:
            current_map = {s["survey_point_name"]: s for s in result["cross_sections_current"]}
            planned_map = {s["survey_point_name"]: s for s in result["cross_sections_planned"]}

            merged = []
            for name, current in current_map.items():
                if name in planned_map:
                    planned = planned_map[name]
                    # 計画高を上書き
                    for i, row in enumerate(current["survey_data"]):
                        if i < len(planned["survey_data"]):
                            row["planned_height"] = planned["survey_data"][i]["elevation"]
                            row["cutting_bottom"] = row["planned_height"] - 0.05
                merged.append(current)

            result["cross_sections_merged"] = merged

    return result


def to_cross_section_ui_format(data: dict) -> list[dict]:
    """cross-section-ui用のフォーマットに変換"""
    merged = data.get("cross_sections_merged", data.get("cross_sections_current", []))

    sections = []
    for s in merged:
        sections.append({
            "survey_point_name": s["survey_point_name"],
            "dl": s["dl"],
            "cl_index": s["cl_index"],
            "l_to_cl_distance": s["l_to_cl_distance"],
            "survey_data": s["survey_data"],
            "route_distance": s["route_distance"],
            "route_id": s.get("route_id", "route_1"),
        })

    return sections


def main():
    if len(sys.argv) < 2:
        xlsx_path = Path(__file__).parent.parent / "data" / "計画.xlsx"
        output_path = xlsx_path.parent / "sections.json"
    else:
        xlsx_path = Path(sys.argv[1])
        output_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else xlsx_path.parent / "sections.json"
    use_gemini = "--local-only" not in sys.argv
    model = "gemini-2.0-flash"
    for i, arg in enumerate(sys.argv):
        if arg == "--model" and i + 1 < len(sys.argv):
            model = sys.argv[i + 1]

    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing: {xlsx_path}", file=sys.stderr)

    if use_gemini:
        try:
            sections = parse_xlsx_with_gemini(xlsx_path, model)
        except Exception as e:
            print(f"Gemini error: {e}", file=sys.stderr)
            sys.exit(1)
    else:
        data = parse_xlsx(str(xlsx_path))
        # cross-section-ui用に変換
        sections = to_cross_section_ui_format(data)

    # JSON出力（デフォルトはdata/sections.json）
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(sections, f, ensure_ascii=False, indent=2)

    print(f"Output: {output_path}", file=sys.stderr)
    print(f"Sections: {len(sections)}", file=sys.stderr)

    # 確認用にサマリー出力
    for s in sections[:5]:
        print(f"  {s['survey_point_name']}: {len(s['survey_data'])} points, route={s['route_distance']}m", file=sys.stderr)
    if len(sections) > 5:
        print(f"  ... and {len(sections) - 5} more", file=sys.stderr)


if __name__ == "__main__":
    main()
